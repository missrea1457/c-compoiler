from openpyxl import load_workbook
def action(s,a):
    actionlist = ['0', 'State', 'WHITESPACE', '*', '+', '-', '/',
              '%', '^', '+=', '-=', '++', '--', '<', '<=', '>',
              '>=', '=', '==', '!=', '||', '&&', '~', 'main', 'id',
              'num', 'string', 'character', 'int', 'char', 'float',
              'double', 'void', 'cin', 'cout', 'while', 'for', 'if',
              'else', 'break', 'return', '[', ']', '{', '}', '(',
              ')', ',', ';', '<<<', '>>>', '>>', '<<', '$']
    workbook = load_workbook(filename="act.xlsx")
    act = workbook.active
    col = actionlist.index(a)
    return act.cell(row=int(s)+2, column=col).value
def goto(t , A):
    gotolist = ['0', 'State', "MAIN'", 'MAIN1', 'F',
            'MAIN', 'STM', 'S', 'A', 'W1', 'C0', 'E',
            'H', 'H1', 'I', 'K', 'J', 'A1', 'A2', 'A3', 'A12',
            'A4', 'A5', 'A8', 'A9', 'A6', 'A7', 'B', 'A16',
            'A15', 'A10', 'A11', 'A13', 'A14', 'B1', 'B2', 'B6',
            'B5', 'B3', 'B9', 'B10', 'C', 'C1', 'C2', 'L1', 'L',
            'L2', 'L3', 'L4', 'L5', 'E3', 'E4', 'E1', 'E5', 'T',
            'N', 'W2', 'W4', 'MS', 'US', 'MS1', 'K1', 'K2', 'K3',
            'K4', 'H2', 'F12', 'F13', 'F14']
    workbook = load_workbook(filename="goto.xlsx")
    goto = workbook.active
    cul =gotolist.index(A)
    return goto.cell(row=int(t)+2, column=cul).value
def reduce(number):
    list = [
        ["MAIN\'",1], #0
        ['MAIN', 2],  # 1
        ['MAIN1', 6],  # 2
        ['STM', 1],  # 3
        ['S', 2],  # 4
        ['S', 2],  # 5
        ['S', 2],  # 6
        ['S', 2],  # 7
        ['S', 2],  # 8
        ['S', 2],  # 9
        ['S', 2],  # 10
        ['S', 2],  # 11
        ['S', 1],  # 12
        ['S', 0],  # 13
        ['A', 1],  # 14
        ['A1', 2],  # 15
        ['A1', 2],  # 16
        ['A1', 2],  # 17
        ['A2', 1],  # 18
        ['A2', 1],  # 19
        ['A2', 1],  # 20
        ['A2', 1],  # 21
        ['A3', 2],  # 22
        ['A4', 1],  # 23
        ['A4', 4],  # 24
        ['A5', 3],  # 25
        ['A5', 1],  # 26
        ['A6', 1],  # 27
        ['A6', 3],  # 28
        ['A7', 1],  # 29
        ['A7', 1],  # 30
        ['A15', 2],  # 31
        ['A15', 2],  # 32
        ['A15', 2],  # 33
        ['A16', 3],  # 34
        ['A16', 3],  # 35
        ['A16', 1],  # 36
        ['A8', 1],  # 37
        ['A9', 1],  # 38
        ['A9', 3],  # 38
        ['A9', 3],  # 39
        ['A10', 2],  # 40
        ['A10', 2],  # 41
        ['A11', 3],  # 42
        ['A11', 3],  # 43
        ['A11', 2],  # 44
        ['A12', 3],  # 45
        ['A13', 3],  # 46
        ['A13', 2],  # 47
        ['A14', 3],  # 48
        ['A14', 2],  # 49
        ['B', 1],  # 50
        ['B1', 3],  # 51
        ['B1', 3],  # 52
        ['B1', 1],  # 53
        ['B2', 3],  # 54
        ['B2', 3],  # 55
        ['B2', 3],  # 56
        ['B2', 1],  # 57
        ['B6', 3],  # 58
        ['B6', 1],  # 59
        ['B5', 2],  # 60
        ['B5', 2],  # 61
        ['B5', 2],  # 62
        ['B5', 2],  # 63
        ['B5', 2],  # 64
        ['B5', 4],  # 65
        ['B5', 3],  # 66
        ['B5', 1],  # 67
        ['B3', 1],  # 68
        ['B3', 1],  # 69
        ['B3', 1],
        ['B9', 3],  # 70
        ['B10', 2],  # 71
        ['B10', 2],  # 72
        ['B10', 2],  # 73
        ['B10', 2],  # 74
        ['C0', 1],  # 75
        ['C', 2],  # 76
        ['C', 2],  # 77
        ['C1', 3],  # 78
        ['C1', 1],  # 79
        ['C2', 3],  # 80
        ['C2', 3],  # 81
        ['C2', 1],  # 82
        ['L', 1],  # 83
        ['L1', 3],  # 84
        ['L1', 1],  # 85
        ['L2', 3],  # 86
        ['L2', 1],  # 87
        ['L3', 3],  # 88
        ['L3', 3],  # 89
        ['L3', 1],  # 90
        ['L4', 3],  # 91
        ['L4', 3],  # 92
        ['L4', 3],  # 93
        ['L4', 3],  # 94
        ['L4', 1],  # 95
        ['L4', 2],  # 96
        ['L5', 3],  # 97
        ['L5', 1],  # 98
        ['L5', 1],  # 99
        ['E', 2],  # 100
        ['E', 3],  # 101
        ['E', 3],  # 102
        ['E3', 3],  # 103
        ['E3', 6],  # 104
        ['E4', 3],  # 105
        ['E4', 6],  # 106
        ['E4', 1],  # 107
        ['E1', 1],  # 108
        ['E5', 3],  # 109
        ['E5', 3],  # 110
        ['N', 3],  # 111
        ['T', 2],  # 112
        ['T', 2],  # 113
        ['T', 2],  # 114
        ['T', 2],  # 115
        ['T', 2],  # 116
        ['T', 2],  # 117
        ['T', 2],  # 118
        ['T', 2],  # 119
        ['T', 2],  # 120
        ['T', 0],  # 121
        ['J', 2],  # 122
        ['W1', 5],  # 123
        ['W2', 1],  # 124
        ['W2', 1],  # 125
        ['W4', 1],  # 126
        ['W4', 1],  # 127
        ['W4', 1],  # 128
        ['W4', 1],  # 129
        ['W4', 1],
        ['W4', 1],  # 130
        ['W4', 1],  # 131
        ['I', 1],  # 132
        ['I', 1],  # 133
        ['MS', 7],  # 134
        ['MS1', 1],  # 135
        ['MS1', 1],  # 136
        ['MS1', 1],  # 137
        ['MS1', 1],  # 138
        ['MS1', 1],  # 139
        ['MS1', 1],  # 140
        ['MS1', 1],  # 141
        ['MS1', 1],  # 142
        ['MS1', 1],  # 143
        ['US', 5],  # 144
        ['US', 5],  # 145
        ['US', 7],  # 146
        ['K', 9],  # 147
        ['K1', 4],  # 148
        ['K1', 3],  # 149
        ['K1', 0],  # 150
        ['K2', 1],  # 151
        ['K2', 0],  # 152
        ['K3', 3],  # 153
        ['K3', 1],  # 154
        ['K3', 1],  # 155
        ['K3', 0],  # 156
        ['K4', 1],  # 157
        ['K4', 1],  # 159
        ['K4', 1],  # 160
        ['K4', 1],  # 161
        ['K4', 1],  # 162
        ['K4', 1],  # 163
        ['K4', 1],  # 164
        ['K4', 1],  # 165
        ['K4', 1],  # 166
        ['H', 2],  # 167
        ['H1', 2],  # 168
        ['H1', 2],  # 169
        ['H2', 2],  # 170
        ['H2', 2],  # 171
        ['H2', 2],  # 172
        ['F', 4],  # 173
        ['F', 4],  # 174
        ['F', 0],  # 175
        ['F12', 3],  # 176
        ['F13', 2],  # 177
        ['F13', 3],  # 178
        ['F14', 4],  # 179
        ['F14', 1]  # 180

    ]
    return list[number]
tokens=[]
f=open("token.txt", "r")
token=f.readline()
while(token!=''):
    tokens.append(token.strip())
    token=f.readline()
tokens.append("$")
print(tokens)
#get token and save in list tokens
numberoftoken=0
state_stack=["0"]
a=tokens[numberoftoken]
numberoftoken+=1
while(len(state_stack)):
    print(a)
    s=state_stack[len(state_stack)-1]
    prev =action(s,a)
    if prev!=None:
    #let s be top op stack
        if prev[0]=='s':
            print(f"shift state:{prev[1:]}")
            state_stack.append(prev[1:])
            a = tokens[numberoftoken]
            numberoftoken += 1
        elif prev[0]=='r':
            print(f"reduce by rule :{prev[1:]}")
            mpop= reduce(int(prev[1:]))
            for i in range(mpop[1]):
                state_stack.pop()
            t = state_stack[len(state_stack)-1]
            state_stack.append(goto(t,mpop[0]))
        elif prev == 'accept':
            print("Accept")
            break
    else:
        print("error")
        break
